from django.shortcuts import render, redirect
from rest_framework import viewsets
from .serializers import MaquinaSerializer, TurnoSerializer, ProducaoDiariaSerializer
from .forms import ProducaoDiariaForm
from django.db.models import Q, Sum
from django.utils.dateparse import parse_date
from django.shortcuts import render
from .models import ProducaoDiaria, Maquina, Turno
from collections import defaultdict
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.shortcuts import render
from .models import ProducaoDiaria, Maquina
from django.db.models import Sum
from collections import defaultdict


class MaquinaViewSet(viewsets.ModelViewSet):
    queryset = Maquina.objects.all()
    serializer_class = MaquinaSerializer

class TurnoViewSet(viewsets.ModelViewSet):
    queryset = Turno.objects.all()
    serializer_class = TurnoSerializer

class ProducaoDiariaViewSet(viewsets.ModelViewSet):
    queryset = ProducaoDiaria.objects.all().order_by('-data')
    serializer_class = ProducaoDiariaSerializer




def cadastro_producao(request):
    if request.method == 'POST':
        form = ProducaoDiariaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cadastro')
    else:
        form = ProducaoDiariaForm()
    return render(request, 'producao/cadastro.html', {'form': form})




def listar_producao(request):
    data = request.GET.get("data")
    codigo_maquina = request.GET.get("maquina")

    producoes = ProducaoDiaria.objects.all()

    if data:
        producoes = producoes.filter(data=data)
    if codigo_maquina:
        producoes = producoes.filter(maquina__codigo=codigo_maquina)

    agrupado_por_turno = defaultdict(list)
    totais_turno = defaultdict(int)

    for producao in producoes:
        turno_num = producao.turno.numero
        agrupado_por_turno[turno_num].append(producao)
        totais_turno[turno_num] += producao.total_produzido

    total_geral = producoes.aggregate(Sum("total_produzido"))["total_produzido__sum"] or 0

    maquinas = Maquina.objects.all()

    context = {
        "agrupado_por_turno": dict(agrupado_por_turno),  # <-- converte para dict normal
        "totais_turno": dict(totais_turno),  # <-- converte para dict normal
        "total_geral": total_geral,
        "data": data,
        "codigo_maquina": codigo_maquina,
        "maquinas": maquinas
    }

    return render(request, "producao/listagem.html", context)


def home(request):
    if request.method == 'POST':
        form = ProducaoDiariaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listagem')
    else:
        form = ProducaoDiariaForm()
    return render(request, 'producao/home.html', {'form': form})




from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from .models import ProducaoDiaria
from django.utils.dateparse import parse_date


def exportar_pdf(request):
    data = request.GET.get("data")
    codigo_maquina = request.GET.get("maquina")

    producoes = ProducaoDiaria.objects.all()

    # Valida e converte data
    if data:
        try:
            data_formatada = parse_date(data)
            producoes = producoes.filter(data=data_formatada)
        except Exception as e:
            return HttpResponse("Data inválida", status=400)

    if codigo_maquina:
        producoes = producoes.filter(maquina__codigo=codigo_maquina)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_producao.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []

    # Título
    styles = getSampleStyleSheet()
    titulo = Paragraph("Relatório de Produção", styles['Title'])
    elementos.append(titulo)
    elementos.append(Spacer(1, 12))

    # Cabeçalho da tabela
    dados = [["Data", "Máquina", "Operador", "Total", "Código Agulha", "Ocorrências"]]

    # Quebra automática para coluna de ocorrências
    estilo_ocorrencia = styles["BodyText"]
    estilo_ocorrencia.wordWrap = "CJK"

    for p in producoes:
        dados.append([
            p.data.strftime("%d/%m/%Y"),
            p.maquina.codigo,
            p.operador,
            p.total_produzido,
            p.codigo_agulha,
            Paragraph(p.ocorrencias or "", estilo_ocorrencia)
        ])

    # Tabela com larguras ajustadas
    table = Table(dados, colWidths=[70, 60, 80, 50, 70, 180])

    # Estilo da tabela
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-2, -1), 'CENTER'),  # exceto última coluna
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    elementos.append(table)
    doc.build(elementos)
    return response



# View para gerar Excel
def exportar_excel(request):
    data = request.GET.get("data")
    codigo_maquina = request.GET.get("maquina")

    producoes = ProducaoDiaria.objects.all()

    if data and data != "None":
        try:
            data_formatada = datetime.strptime(data, "%Y-%m-%d").date()
            producoes = producoes.filter(data=data_formatada)
        except ValueError:
            pass

    if codigo_maquina and codigo_maquina != "None":
        producoes = producoes.filter(maquina__codigo=codigo_maquina)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produção Diária"

    headers = ["Data", "Máquina", "Operador", "Total Produzido", "Código Agulha", "Ocorrências"]
    ws.append(headers)

    for p in producoes:
        ws.append([
            p.data.strftime("%d/%m/%Y"),
            p.maquina.codigo,
            p.operador,
            p.total_produzido,
            p.codigo_agulha,
            p.ocorrencias
        ])

    # Ajusta largura das colunas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_producao.xlsx'
    wb.save(response)
    return response



